#!/usr/bin/env python3
# coding=UTF-8
# author: hellojing
# date: 2023-01-06
import re,os
import datetime
import argparse,json
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import Border, Side, colors

class HostsCheckTable(object):
    def __init__(self,filepath):
        nowtime = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        basepath = os.getcwd()
        if filepath.endswith("/"):
            self.DestHostsFile = filepath + "HostsCheckTable{0}.xlsx".format(nowtime)
        else:
            self.DestHostsFile = filepath + "/HostsCheckTable{0}.xlsx".format(nowtime)
        self.MetricItem = {"ip":"","hostname":"","os_pretty_name":"","kernel":"","cpuused":"","cpu_loadavg1":"","cpu_loadavg5":"","MemFree":"","MemTotalSize":"","MemAvailable":"","memusedperc":"","root_free_disk":"","root_total_disk":"","root_usage_disk":"","data_total_disk":"","data_free_disk":"","data_usage_disk":""}
        self.SourceHostsFile = os.getcwd() + "/tmp/OriginAnsibleFile.txt"
        self.TmpHostsFile = os.getcwd() + "/tmp/TmpHostExcel.xlsx"

    def GetOriginData(self):
        f1 = json.loads(open(self.SourceHostsFile,"r").read())
        result = f1["plays"][0]["tasks"][0]["hosts"]
        return json.loads(json.dumps(result))

    def GetWriteExcel(self):
        hosts = self.GetOriginData()
        excel = openpyxl.Workbook() # 创建文档
        sheet = excel.create_sheet('hosts', 0) # 创建sheet
        Tabletitle = ['IP地址','主机名','系统信息','内核信息','CPU使用率','Cpu1m负载','Cpu5m负载','内存总空间','内存剩余空间','内存可用空间','内存使用率','/目录总空间','/目录剩余空间','/目录空间使用率','/data/目录总空间','/data/目录剩余空间','/data/目录空间使用率']
        for col in range(len(Tabletitle)):
            c = col + 1
            sheet.cell(row=1, column=c).value = Tabletitle[col]
        all_host_info_list = []
        for ip,v in hosts.items():
            self.MetricItem["ip"] = ip
            try:
                hostinfo = json.loads(v["stdout"])
            except:
                pass
            else:
                self.MetricItem["hostname"] = hostinfo["system"]["hostname"]
                self.MetricItem["os_pretty_name"] = hostinfo["system"]["os_pretty_name"]
                self.MetricItem["kernel"] = hostinfo["system"]["kernel"]
                self.MetricItem["cpuused"] = hostinfo["cpu"]["cpuused"]
                self.MetricItem["cpu_loadavg1"] = hostinfo["cpu"]["cpu_loadavg1"]
                self.MetricItem["cpu_loadavg5"] = hostinfo["cpu"]["cpu_loadavg5"]
                self.MetricItem["MemTotalSize"] = hostinfo["mem"]["MemTotalSize"]
                self.MetricItem["MemFree"] = hostinfo["mem"]["MemFree"]
                self.MetricItem["MemAvailable"] = hostinfo["mem"]["Available"]
                self.MetricItem["memusedperc"] = hostinfo["mem"]["usedperc"]
                self.MetricItem["root_total_disk"] = hostinfo["disk"]["root_total_disk"]
                self.MetricItem["root_free_disk"] = hostinfo["disk"]["root_free_disk"]
                self.MetricItem["root_usage_disk"] = hostinfo["disk"]["root_usage_disk"]
                self.MetricItem["data_total_disk"] = hostinfo["disk"]["data_total_disk"]
                self.MetricItem["data_free_disk"] = hostinfo["disk"]["data_free_disk"]
                self.MetricItem["data_usage_disk"] = hostinfo["disk"]["data_usage_disk"]
            # 设置有序的指标列表
            col_value_list = [self.MetricItem["ip"] ,self.MetricItem["hostname"],self.MetricItem["os_pretty_name"],self.MetricItem["kernel"],self.MetricItem["cpuused"],self.MetricItem["cpu_loadavg1"],self.MetricItem["cpu_loadavg5"],self.MetricItem["MemTotalSize"],self.MetricItem["MemFree"],self.MetricItem["MemAvailable"],self.MetricItem["memusedperc"],self.MetricItem["root_total_disk"],self.MetricItem["root_free_disk"],self.MetricItem["root_usage_disk"],self.MetricItem["data_total_disk"],self.MetricItem["data_free_disk"],self.MetricItem["data_usage_disk"]]
            all_host_info_list.append(col_value_list)
        for row in range(len(all_host_info_list)):
            sheet.append(all_host_info_list[row])
        excel.save(self.TmpHostsFile)
        wb = openpyxl.load_workbook(self.TmpHostsFile)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        max_row = ws.max_row
        max_column = ws.max_column
        align=Alignment(horizontal='center',vertical='center')
        border = Border(left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'))
        for i in range(1, max_row + 1):
            for j in range(1, max_column + 1):
                if i == 1:
                    ws.cell(i,j).font = Font(u'宋体',size = 10,bold=True,strike=False,color='000000')
                ws.cell(i, j).alignment = align
                ws.cell(i, j).border = border
        wb.save(self.DestHostsFile)
        print("主机巡检报表{0}已生成。".format(self.DestHostsFile))
        DelTmpFile = self.DelTmpFile()
    def DelTmpFile(self):
        if (os.path.isfile(self.TmpHostsFile)):
            os.remove(self.TmpHostsFile)
        if (os.path.isfile(self.SourceHostsFile)):
            os.remove(self.SourceHostsFile)

if __name__ == '__main__':
    DefaultDir = os.getcwd() + "/result"
    if not (os.path.exists(DefaultDir)):
        os.mkdir(DefaultDir)
    parser = argparse.ArgumentParser()
    parser.add_argument("-p", "--PATH",default=DefaultDir , help="添加输出路径地址.")
    args = parser.parse_args()
    HostsCheckExcelPath = args.PATH
    HostsCheckExcel = HostsCheckTable(HostsCheckExcelPath).GetWriteExcel()
